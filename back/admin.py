from django.contrib import admin
from django.contrib.auth.models import Group
from django.contrib import messages
from django.utils.translation import ngettext
from rest_framework.authtoken.models import TokenProxy
from django.http import HttpResponse
from rest_framework.authtoken.models import Token
# Create your views here.
from django.contrib.auth.tokens import default_token_generator
from django.shortcuts import redirect, render
from django.http import HttpResponseRedirect
from .email import send_html_mail
from .models import *
import re
from main.models import *
from openpyxl import Workbook
from django.utils.html import format_html
from openpyxl.utils import get_column_letter
import datetime
from django.contrib.admin.views.main import ChangeList
from django.contrib.admin.utils import quote

admin.site.unregister(TokenProxy)
admin.site.unregister(Group)












@admin.action(description='Export User Details to XLSX')
def export(self, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=User Details '+str(len(queryset))+'.xlsx'
    wb=Workbook()
    ws=wb.active
    ws.title='Users'
    row_num = 1
    columns=['Username','Email','Name','Contact','Date Joined','Total Reputation','Email Confirmed']
    # print(len(columns))
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
    rows = queryset
    for ro in rows:
        row_num += 1
        row = [ro.username,ro.email,str(ro.name),str(ro.contact),str(ro.date_joined),str(ro.totalreputation),'Yes' if ro.email_confirmed else 'No']
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)
    wb.save(response)
    return response

@admin.action(description='Export Emails only to XLSX')
def exportemails(self, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=User Details '+str(len(queryset))+'.xlsx'
    wb=Workbook()
    ws=wb.active
    ws.title='Users'
    row_num = 1
    columns=['Email']
  #  print(len(columns))
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
    rows = queryset
    for ro in rows:
        row_num += 1
        row = [ro.email]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)
    wb.save(response)
    return response


























# def third_largest(arr): 
#     three_largest = [None,None,None] 
#     for el in arr: 
#         if three_largest[2]!=None and el < three_largest[2]: 
#             #skip early if not a candidate 
#             continue 
#         for i in range(3): 
#             #Check against largest first: 
#         	if three_largest[i] == None or el > three_largest[i]: 
#                 three_largest[i] = el 
#                 break 
#     return three_largest[2]




class UserAdmin(admin.ModelAdmin):
    # view_on_site = False
    date_hierarchy = 'date_joined'
    # def get_changelist(self, request, **kwargs):
    #     return UserChangeList
    # def has_change_permission(self, request, obj=None):
    #     if obj:
    #         if request.user.id == obj.id:
    #             return True
    #         else:
    #             return False
    #     else:
    #         return False
    
    def has_delete_permission(self, request, obj=None):
        return request.user.is_superadmin  # type: ignore
        
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        response = super(UserAdmin, self).changelist_view(request, extra_context)
        # print(response)
        extra_context['date_name'] = 'Date Joined'
        if request.method == 'GET':
            try:
                res = response.context_data["cl"].queryset
                card2 = []
                for card in res:
                    # print(card)
                    card2.append({'primary': card.name, 'secondary': card.username+ ' • '+str(card.totalreputation), 
                    'image': card.photo.url if card.photo else '/static/user.jpg',
                    'type': card.get_type_display(), 'ok': card.email_confirmed,
                    'actions': ['<a href="sendemail">Email</a>']})
                extra_context['cards'] = card2  # type: ignore
            except Exception as e:
                print(e)
        response = super(UserAdmin, self).changelist_view(request, extra_context)

        return response

    #URLS
    def get_urls(self):
            from django.urls import path
            urls2 = [
                path('sendemailadmin/', self.sendemailadmin, name='sendemailadmin'),
                path('make_staff/', self.makestaff, name='make_staff'),
                path('modify-reputation/', self.modifyrep, name='modify-reputation'),
                path('profile/', self.profile, name='profile'),
            ]
            urls = super().get_urls()
            # print(urls)
            urls2 += urls
            return urls2

    #VIEWS
    def sendemailadmin(self, request, **kwargs):
        userids = request.GET['ids'].split(',')
        files = []
        for file in request.FILES.getlist('files'):
            files.append({'name': file.name, 'file': file.read(),'type': file.content_type})
        useremails = []
        for id in userids:
            user = User.objects.get(id=id)
            useremails.append(user.email)
        send_html_mail(request.POST['subject'],
        {'title': request.POST['title'],'content': request.POST['content'],'remark': request.POST['remark'] if 'remark' in request.POST else None,'link': {'title': request.POST['linkname'],'action': request.POST['linktarget']}},
        request.POST['template'],useremails,[],files,[])
        messages.success(request, "Email was successfully sent to {} Users".format(str(len(userids))))
        return redirect('/panel/back/user')
        

    def makestaff(self, request, **kwargs):
        userids = request.GET['ids'].split(',')
        useremails = []
        company = User.objects.get(id=request.POST['company'])
        for id in userids:
            user = User.objects.get(id=id)
            user.type = 'Ct'
            user.refuser = company
            user.save()
            useremails.append(user.email)
        send_html_mail('Marked Company Staff', {'title': 'Marked Company Staff','content': 'You were made as Company Saff to '+company.name,'remark': 'Beware of Spams'},'email/send.html',useremails,[],[],[])
        messages.success(request, "{} Users were successfully made Staff".format(str(len(userids))))
        return redirect('/panel/back/user')

    def modifyrep(self, request, **kwargs):
      #  print(request.method)
        # if request.method == 'GET':
        #     return redirect('/panel/back/user/?q=vivek')
        users = request.POST.getlist('users')
      #  print(request.POST)
      #  print('entered')
        for u in users:
            user = User.objects.get(id=u)
            user.refreputation += int(request.POST['refreputation'])
            user.upvotereputation += int(request.POST['upvotereputation'])
            user.likesreputation += int(request.POST['likesreputation'])
            user.verificationreputation += int(request.POST['verificationreputation'])
            user.save()
            #mail here
        messages.success(request, "Reputation of {} Users was modified".format(str(len(users))))
        return redirect('/panel/back/user/')


        return HttpResponseRedirect('/panel/back/user')

    def profile(self, request, **kwargs):
        return HttpResponseRedirect('/panel/back/user')

    #OPTIONS
    list_display = ('userv','email','name','atype','totalreputation','refcode','refs','date_joined') #,'is_superadmin','is_admin','is_company','is_staff','refreputation','upvotereputation','likesreputation','verificationreputation'     
    list_filter = ('type', 'email_confirmed','date_joined','is_active')
    search_fields = ['username', 'email', 'name','refcode', 'refuser__username','refuser__name','refuser__email']
    # 
    actions = [ 'sendemail','make_admin','make_company',export,exportemails,'make_staff','make_user','freeze','unfreeze','modify','sendverificationemail']
    exclude = ('password','email','type','contact','email_confirmed','refcode','phone_confirmed','is_active','reset','photo','is_superadmin','is_admin','is_company','is_staff','visits','refreputation','upvotereputation','likesreputation','verificationreputation','otherreputation','totalreputation')
    autocomplete_fields = ['refuser']
    readonly_fields = ['userv','atype','refs','website', 'reputation']
    
    def reputation(self, obj):
        if obj.type == 'C' or obj.type == 'Ct':
            return False
        else:
            return format_html('''
            <span class="mr-2 py-1 px-2 rounded-circle text-white bg-primary" rel="tooltip" title="Referral Reputation">%s</span>+
            <span class="mr-2 py-1 px-2 rounded-circle text-white bg-info" rel="tooltip" title="Likes Reputation">%s</span>+
            <span class="mr-2 py-1 px-2 rounded-circle text-white bg-error" rel="tooltip" title="Upvotes Reputation">%s</span>+
            <span class="mr-2 py-1 px-2 rounded-circle text-white bg-warning" rel="tooltip" title="Verification Reputation">%s</span>+
            <span class="mr-2 py-1 px-2 rounded-circle text-whitecs bg-secondary" rel="tooltip" title="Other Reputation">%s</span>=
            <span class="mr-2 py-1 px-2 rounded-circle text-dark bg-success" rel="tooltip" title="Total Reputation">%s</span>
            '''%(str(obj.refreputation),str(obj.likesreputation),str(obj.upvotereputation),str(obj.verificationreputation),str(obj.otherreputation),str(obj.totalreputation)))

    def get_actions(self, request):
        actions = super(UserAdmin, self).get_actions(request)
        if not request.user.is_superadmin:
            del actions['make_user']
            del actions['make_company']
            del actions['make_admin']
            del actions['make_staff']
            del actions['freeze']
            del actions['unfreeze']
            # [actions['make_user'],actions['make_company'],actions['make_staff'],actions['make_admin'],actions['freeze'],actions['unfreeze']]
        else:
            pass
        return actions

    #HTML FORMATTING
    def refs(self, obj):
        if obj.refuser:
            return format_html('<a href="%s/change">%s</a>'%(obj.refuser.id,obj.refuser.username))
        else:
            return '-'
    def atype(self, obj):
        return format_html('<div class="user-type bg-type-%s" rel="tooltip" data-placement="left" title="%s"></div>'%(obj.get_type_display(),obj.get_type_display()))

    def userv(self, obj):
        return format_html('<div class="user-link "><span class="%s" rel="tooltip" title="%s"></span><img src="%s" width="25" height="25" style="width: 25px !important;" ><p>%s</p>%s</div>'%('active' if obj.is_active else 'inactive','Active' if obj.is_active else 'Freezed',
            obj.photo.url if obj.photo else '/static/user.jpg','' if not re.match("^[a-zA-Z0-9_@ .-]+$", obj.username) else obj.username,'<img src="/static/admin/img/icon-yes.svg" alt="Verified" rel="tooltip" title="Verified" class="ml-1">' if obj.email_confirmed else ''))

    userv.short_description = 'User'
    userv.admin_order_field = 'username'
    refs.short_description = "Attached To"
    refs.admin_order_field = 'refuser__username'
    atype.short_description = 'Type'
    atype.admin_order_field = 'type'

    #ACTIONS
    def sendverificationemail(self, request, queryset):
        if not request.user.is_authenticated:
            return HttpResponseRedirect('/panel/login/?next='+request.path) 
        if request.method == 'POST':
            for q in queryset:
                user = User.objects.get(id=q.id)
                tok = Token.objects.create(user=user)
                token = default_token_generator.make_token(user)
                send_html_mail('Activate Your Account ',{'link': 'https://app.cyber3ra.com/verify/?token=' + str(tok.key)+'&user='+ token,'name': user.name if user.name else user.username,'code': token},'email/verify2.html',[user.email],[],[],[])
                self.message_user(request, ngettext('%d user was sent a Verification Email', '%d users were sent a Verification Email', len(queryset), ) % len(queryset), messages.SUCCESS)
    def sendemail(self, request, queryset):
        if not request.user.is_authenticated:
            return HttpResponseRedirect('/panel/login/?next='+request.path) 
        if request.method == 'POST':
            return render(request, 'custom/emails.html', context={'users':queryset,'custom': 'Yes','link': 'sendemailadmin/?ids=%s' % (
                ','.join(str(q.id) for q in queryset),
    )})
    def modify(self, request, queryset):
        if not request.user.is_authenticated:
            return HttpResponseRedirect('/panel/login/?next='+request.path) 
      #  print(request.META.get('HTTP_REFERER', '/'))
        return render(request,'custom/reputation.html',{'users':queryset,'link': 'modify-reputation/?ids=%s' % (
                ','.join(str(q.id) for q in queryset))})
    def make_admin(self, request, queryset):
        updated = queryset.update(type='A')
        useremails = []
        for query in queryset:
            user = User.objects.get(id=query.id)
            useremails.append(user.email)
        send_html_mail('Marked Admin', {'title': 'Marked Admin','content': 'You were marked as Admin','remark': 'Beware of Spams'},'email/send.html',useremails,[],[],[])
        self.message_user(request, ngettext('%d user was successfully made as Admin', '%d users were successfully made Admin', updated, ) % updated, messages.SUCCESS)
    def make_company(self, request, queryset):
        updated = queryset.update(type='C')
        useremails = []
        for query in queryset:
            user = User.objects.get(id=query.id)
            useremails.append(user.email)
        send_html_mail('Marked Company', {'title': 'Marked Company','content': 'You were marked as Company','remark': 'Beware of Spams'},'email/send.html',useremails,[],[],[])
        self.message_user(request, ngettext('%d user was successfully made as Company', '%d users were successfully made Company', updated, ) % updated, messages.SUCCESS)
    def make_staff(self, request, queryset):
        if not request.user.is_authenticated:
            return HttpResponseRedirect('/panel/login/?next='+request.path) 
        if request.method == 'POST':
            return render(request, 'custom/make_staff.html', context={'users':User.objects.filter(type='C'),'link': 'make_staff/?ids=%s' % (
                ','.join(str(q.id) for q in queryset),
        )})
    def make_user(self, request, queryset):
        updated = queryset.update(type='U')
        useremails = []
        for query in queryset:
            user = User.objects.get(id=query.id)
            useremails.append(user.email)
        send_html_mail('Marked User', {'title': 'Marked User','content': 'You were marked as User','remark': 'Beware of Spams'},'email/send.html',useremails,[],[],[])
        self.message_user(request, ngettext('%d user was successfully made as User', '%d users were successfully made User', updated, ) % updated, messages.SUCCESS)
    def freeze(self, request, queryset):
        updated = queryset.update(is_active=False)
        useremails = []
        for query in queryset:
            user = User.objects.get(id=query.id)
            useremails.append(user.email)
        send_html_mail('Account Freezed', {'title': 'Account Freezed','content': 'Your account was freezed.','remark': 'Beware of Spams'},'email/send.html',useremails,[],[],[])
        self.message_user(request, ngettext('%d account was Freezed', '%d accounts were Freezed', updated, ) % updated, messages.WARNING)
    def unfreeze(self, request, queryset):
        updated = queryset.update(is_active=True)
        useremails = []
        for query in queryset:
            user = User.objects.get(id=query.id)
            useremails.append(user.email)
        send_html_mail('Account Un-Freezed', {'title': 'Account Un-Freezed','content': 'Your account was un-freezed.','remark': 'Beware of Spams'},'email/send.html',useremails,[],[],[])
        self.message_user(request, ngettext('%d account was Un-Freezed', '%d accounts were Un-Freezed', updated, ) % updated, messages.SUCCESS)
    # def rfreeze(self, request, queryset):
    #     updated = queryset.update(is_active=False)
    #     self.message_user(request, ngettext('%d account was Freezed', '%d accounts were Freezed', updated, ) % updated, messages.WARNING)
    # def runfreeze(self, request, queryset):
    #     updated = queryset.update(is_active=True)
    #     self.message_user(request, ngettext('%d account was Un-Freezed', '%d accounts were Un-Freezed', updated, ) % updated, messages.SUCCESS)
    sendemail.short_description = "Send Email to Selected Users"
    sendverificationemail.short_description = "Send Verification Email to Selected Users"
    make_admin.short_description = "Make Selected Users Admin"
    make_company.short_description = "Make Selected Users Company"
    make_staff.short_description = "Make Selected Users Company Staff"
    make_user.short_description = "Make Selected Users User"
    freeze.short_description = "Freeze Selected Accounts"
    unfreeze.short_description = "Liquefy Selected Accounts"   
    modify.short_description = 'Modify Reputation of Selected Users'
    # def __init__(self,*args,**kwargs):
    #     super(UserAdmin, self).__init__(*args, **kwargs)
    #     self.list_display_links = (None, )


admin.site.register(User,UserAdmin)
admin.site.site_header = 'Cyber3ra Administration'
admin.site.site_title = 'Cyber3ra Administration'


