#ADMIN VIEWS
from decimal import Decimal
from http.client import HTTPResponse
from traceback import print_tb

from back.email import send_html_mail
from .models import User
from main.models import Program
from django.shortcuts import redirect, render
from django.contrib import messages
# from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
# from rest_framework.decorators import permission_classes
from django.contrib.auth.models import Permission
from honeb.settings import EMAIL_HOST_NAME
import os
import csv

from django.db.models.query_utils import Q
import openpyxl
from django.core.mail import EmailMessage

import sys

from django.core.management import call_command







def adminprofile(request):
    # print(request)
    if not request.user.is_authenticated:
        return redirect('/panel/login/?next='+request.path) 
    if request.method == 'GET':
        profile = 35 if request.user.name else 20
        profile+=20 if request.user.photo else 0
        profile+=15  if request.user.email_confirmed else 0
        all = [
            {'label': 'Username (10%)','val': True},
            {'label': 'Email (10%)','val': True},
            {'label': 'Full Name (15%)','val': True if request.user.name else False},
            {'label': 'Profile Picture (20%)' if request.user.is_admin else 'Company Logo (20%)' ,'val': True if request.user.photo else False},
            {'label': 'Email Confirmed (15%)','val': request.user.email_confirmed}
        ]

        addedprogram = False
        if request.user.is_admin:
            progs = Program.objects.filter(posted_by=request.user)
            # if len(progs) > 0:
            #     profile += 30
            #     addedprogram = True
            all.append({'label': 'Admin (30%)','val': True})
        else:
            progs = Program.objects.filter(posted_by=request.user)
            if len(progs) > 0:
                profile += 30
                addedprogram = True
            all.append({'label': 'Added Program (30%)','val': addedprogram})
        return render(request,'custom/profile.html',{'complete': profile,'all': all,'programs': progs })
    elif request.method == 'POST':
      #  print(request.POST['bio'])
        request.user.name = request.POST['name']
        request.user.bio = request.POST['bio']
        request.user.save()
        return redirect('/panel/profile')
    else:
        return redirect('/panel/')




def profilepic(request):
	if request.user.is_authenticated:
		if request.method == 'GET':
			request.user.photo.delete()
			return redirect('/panel/profile')
		if request.method == 'POST':
			# print(request.FILES)
			request.user.photo = request.FILES['file']
			request.user.save()
			# print(request.POST)
			return redirect('/panel/profile')
	else:
		return redirect('/panel/login/?next='+request.path)


def sendemail(request):
    # print(request)
    if not request.user.is_authenticated:
        return redirect('/panel/login/?next='+request.path) 
    if not request.user.is_admin:
        return redirect('/panel/') 
    if request.method == 'GET':
        context = {}
        # print(request.scheme)
        context["users"] = User.objects.filter(type='U',email_confirmed=True).values_list('id','name','email','username')
        context["companies"] = User.objects.filter(type='C').values_list('id','name','email','username')
        return render(request,'custom/emails.html',context)
    if request.method == 'POST':
      #  print(request.POST)
        # print(request.FILES.getlist('files'))
      #  print(request.POST.getlist('users'))
        ids = request.POST.getlist('users')
        files = []
        for file in request.FILES.getlist('files'):
            files.append({'name': file.name, 'file': file.read(),'type': file.content_type})
        useremails = []
        for id in ids:
            user = User.objects.get(id=id)
            useremails.append(user.email)
            
        # msg = EmailMessage(request.POST['subject'], request.POST['content'], EMAIL_HOST_NAME, [],useremails)
        # msg.send()
        send_html_mail(request.POST['subject'],
        {'title': request.POST['title'],'content': request.POST['content'],'remark': request.POST['remark'] if 'remark' in request.POST else None,'link': {'title': request.POST['linkname'],'action': request.POST['linktarget']}},
        request.POST['template'],useremails,[],files,[])
        messages.success(request, "Email was successfully sent to {} Users".format(str(len(ids))))
        return redirect('/panel/')



def settings(request):
    # print(request)
    if not request.user.is_authenticated:
        return redirect('/panel/login/?next='+request.path) 
    return render(request,'custom/settings.html')

def about(request):
    # print(request)
    perms = Permission.objects.exclude(content_type__in=['3','2','6','7','4','5'])
    # print(perms)
    # for p in perms:
    #   #  print(p.content_type.id)
    #   #  print(p.content_type.name)
    if not request.user.is_authenticated:
        return redirect('/panel/login/?next='+request.path) 
    return render(request,'custom/about.html',{'permissions':perms})

    


def createbackup(request):
    # print(request)
    if not request.user.is_authenticated:
        return redirect('/panel/login/?next='+request.path) 
    if not request.user.is_superadmin:
        messages.warning(request, "Only Super Admins can do that")
        return redirect('/panel/')
    print("here")
    sysout = sys.stdout
    # call_command('set',)
    sys.stdout = open('filename.xml', 'w')
    call_command('dumpdata',format="xml") 
    sys.stdout = sysout
    return redirect('/panel/about') 







def duplicate(request):
    # print(request)
    if not request.user.is_authenticated:
        return redirect('/panel/login/?next='+request.path) 
    if not request.user.is_superadmin:
        messages.warning(request, "Only Super Admins can do that")
        return redirect('/panel/') 
    path = "C:\\Users\\Vivek Billa\\cyback\\users.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    dictt = {}
    eles = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 0):
        temp = {}
        if(i == 0):
            head = list(row)
        else:
            for j, col in enumerate(row, 0):
                if col != None:
                    temp[head[j].lower()] = col
            # print(temp)
            eles.append(temp)
    # for row in range(1, ws.max_row + 1):
    #     print(row)
    #     key = ws.cell(row, 1).value
    #     value = ws.cell(row, 2).value
    #     dictt[key] = value
    # print(eles)
    num = 0
    for ele in eles:
        # print(ele['username'], ele['email'])
        try:
            user2 = User.objects.get(Q(username__iexact=ele['username'])|Q(email=ele['email']))
            print("Exists", ele['username'])
            num+=1
        except Exception as e:
            try:
                user = User.objects.create_user(username=ele['username'], email=ele['email'], name=ele['name'] if 'name' in ele else '', other=int(ele['total reputation']), active=False, password=ele['username']+"@SecuriumX")
            except Exception as er:
                print(er, ele['username'], ele['email'])
    print(num)
    # return 
    return redirect('/panel/about') 
